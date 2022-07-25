"""
PEP8 - import
Например - os - это стандартная библеотека, поэтому ее правильно поместить в начало.
Пути лучше через os.path.join или pathlib - отдельная библиотека
почить про sql injection
"""
import os
import json
import pandas as pd
import psycopg2
import openpyxl
from PIL import ImageColor
import logging
import configparser
from openpyxl.worksheet.datavalidation import DataValidation

config = configparser.ConfigParser()
config.read("connect_art.ini", encoding='utf-8')

# # Config
host = config.get('DATABASE', 'host')
user = config.get('DATABASE', 'user')
password = config.get('DATABASE', 'password')
database = config.get('DATABASE', 'database')
path_templatecreate = config.get('FILES', 'path_templatecreate')
path_templatecreated = config.get('FILES', 'path_templatecreated')
path_storage_templates = config.get('DEFAULT', 'storage') + "/templates"
path_storage_template_group = config.get('DEFAULT', 'storage') + "/template_group"
path_storage_subjects = config.get('DEFAULT', 'storage') + "/subjects"

# logging
logger = logging.getLogger("log.connect")


# functions
def get_conn():
    """
    Cоздает и возвращает объект соедининия
    :return:
    """
    conn = None
    try:
        conn = psycopg2.connect(
            host=host,
            user=user,
            password=password,
            database=database)
        logger.debug("Connection is successful")

    except Exception as ex:
        logger.error("Error while working with PostgreSQL")
    return conn


def execute_query(connection, query):
    """
    Для вставки, удаления, изменения данных
    :param connection: соединение
    :param query: SQL запрос
    """
    connection.autocommit = True
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        logger.debug("Query is successful")
        cursor.close()
        connection.close()
        logger.debug("Соединения с PostgreSQL закрыто")
        return True
    except Exception as ex:
        logger.error(ex)
        if connection:
            connection.close()
            logger.debug("Соединения с PostgreSQL закрыто")
        return False


def execute_read_query(connection, query):
    """
    Для извлечения данных
    :param connection: соединение
    :param query: SQL запрос
    :return: данные таблицы SQL
    """
    cursor = connection.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        logger.debug("Query is successful")
    except Exception as ex:
        logger.error(ex)
        """connection.close  или with ..."""
    finally:
        if connection:
            cursor.close()
            connection.close()
            logger.debug("Соединения с PostgreSQL закрыто")
        return result


def get_users_templates():
    """
    :return: возращает (генератор: путь, файл)
    плюс проверка директории
    """
    sql = """SELECT coalesce(path_to_home, '') || coalesce(path_to_templates, ''),
          coalesce(path_to_home, '') || coalesce(path_to_templ_conf, '') FROM author;"""
    result = execute_read_query(get_conn(), sql)
    for user, file in result:
        if os.path.exists(file[1:]):
            yield user[1:], file[1:]
        else:
            logger.error(f"Файл {file} не существует")


def to_json(xlsx):
    """
    Функция создает файл json
    :param file_xlxs: Файл xlsx
    :return: Возвращает имя json файла
    """
    path, name = xlsx.rsplit("/", 1)
    json_name = f"{name.split('.')[0]}.json"
    import pandas as pd
    try:
        data = pd.read_excel(xlsx)
        data.to_json(os.path.join(path, json_name), indent=4, orient="records")
    except ValueError:
        with open(os.path.join(path, json_name), 'wt', encoding='utf-8') as file:
            json.dump(file)
    finally:
        return json_name


def get_group(path, name):
    """
    Фукнция возвращает наименование группы (источник файл Templates)
    :param path: путь к папке Templates пользователя
    :param name: имя template
    :return: возвращает группу
    """
    try:
        df = pd.read_excel(path)
        d = {k: v for k, v in zip(df["templates"], df["group"])}
        return d.get(name, None)
    except FileNotFoundError:
        logger.error("файл Templates.xlsx отсутствует")
        return False


def data_validation_temp(workbook, file):
    """
    Функция создает проверку данных в Templates
    :param file: путь к файлу templates
    :param workbook: объект workbook
    """
    # обнуляем на всякий случай проверку
    ws = workbook.active
    ws.data_validations.dataValidation = []
    sql = """select name from template_group"""
    data = f'"{",".join(x[0] for x in execute_read_query(get_conn(), sql))}"'
    dv = DataValidation(type='list', formula1=data, allow_blank=True, errorStyle='warning')
    ws.add_data_validation(dv)
    dv.add("B2:B100")
    workbook.save(file)


def update_template(template, templates_xlsx, template_create, template_created):
    """
    Фукнкция принимает название template пути к template_create, template_created
    и обновляет эти файлы
    """
    try:
        wb = openpyxl.load_workbook(templates_xlsx)
        ws = wb.active
        delete_row = []
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                if template == cell.value:
                    delete_row.append(int(cell.coordinate[-1]))
        for row in delete_row:
            ws.delete_rows(row)
        wb.save(templates_xlsx)
        wb.close()
    except Exception as ex:
        logger.error(f"Проверьте наличие файла {templates_xlsx}")
    try:
        wb = openpyxl.load_workbook(template_created)
        ws = wb.active
        if template not in [cell.value for row in ws.iter_rows(min_row=2, min_col=1) for cell in row]:
            for row in ws.iter_rows(min_row=2, min_col=1):
                if row[0].value is None:
                    row[0].value = template
                    break
            else:
                length = ws.max_row
                ws[f"A{length + 1}"] = template
            wb.save(template_created)
            wb.close()
    except Exception as ex:
        logger.error(f"Проверьте наличие файла {template_created}")
    try:
        wb = openpyxl.load_workbook(template_create)
        ws = wb.active
        delete_row = []
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                if template == cell.value:
                    delete_row.append(int(cell.coordinate[-1]))
        for row in delete_row:
            ws.delete_rows(row)
        wb.save(template_create)
        wb.close()
    except Exception as ex:
        logger.error(f"Проверьте наличие файла {template_create}")


def get_users_subjects():
    """
    :return: генератор: путь, файл
    """
    sql = """SELECT coalesce(path_to_home, '') || coalesce(path_to_subjects, ''), 
    coalesce(path_to_home, '') || coalesce(path_to_subj_conf, ''), name FROM author;"""
    result = execute_read_query(get_conn(), sql)
    for user, file, name in result:
        if os.path.exists(file[1:]):
            yield user[1:], file[1:], name
        else:
            logger.error(f"Файл {file} не существует")


def check_path(subj, user, list_products):
    """
    :param name: имя subjects
    :param user: путь к рабочей директории user
    :return: возращает True если все диретории из списка true_path существуют в рабочей директории user
    """
    sql = """select distinct(path_type) from product_type"""
    result = execute_read_query(get_conn(), sql)
    paths = [path[0] for path in result if path[0].split('/')[1] in list_products]
    folders = ('/prints', '/mockups')
    for path in paths:
        for folder in folders:
            if os.path.exists(user + f"/{subj.value}" + path + folder) is False:
                logger.error(f"У {user}/{subj.value} нет директории {path}")
                return False
    return True


def is_correct(name, user):
    """
    :param name: имя subjects
    :param user: путь к рабочей директории user
    :return: возращает True если все диретории из списка true_path существуют в рабочей директории user
    """
    sql = "select distinct(path_type) from product_type"
    result = execute_read_query(get_conn(), sql)
    paths = [path[0] for path in result]
    folders = ('/prints', '/mockups')
    for path in paths:
        for folder in folders:
            if os.path.exists(user + f"/{name.value}" + path + folder) is False:
                logger.error(f"У {user}/{name.value} нет директории {path}")
                return False
    return True


def template_exist(template):
    """
    проверка наличия шаблона template в БД
    :param template:  название шаблона
    :return: True - если шаблон существует, False если нет
    """
    query = f"""select name from template where name = '{template.value}'"""
    response = execute_read_query(get_conn(), query)
    if not response:
        logger.error(f"Template {template.value} отсутствует в БД")
        return False
    return True


def append_data_validation_subj(name, coordinate, ws):
    """
    Фукнкция для проверки данных в файле Subjects
    :param ws: объект worksheets
    :param name: - название категории
    :param coordinate: столбец Excel
    """
    if name in {'template', 'brand', 'infographic_color', 'image_position', 'common_photo', 'mockup', 'infographic'}:
        sql = f"""select distinct(name) from {name}"""
        data = f'"{",".join(x[0] for x in execute_read_query(get_conn(), sql))}"'
        dv = DataValidation(type="list", formula1=data, allow_blank=True, errorStyle='warning')
        ws.add_data_validation(dv)
        dv.add(f'{coordinate}3:{coordinate}100')
    elif 'product' in name:
        dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True, errorStyle='warning')
        ws.add_data_validation(dv)
        dv.add(f'{coordinate}3:{coordinate}100')


def data_validation_subj(workbook, file):
    """
    Функция работает с файлом Subjects она принимает объект workbook, парсит файл и вызывает функцию
    append_data_validation для установления проверки данных
    :param file: путь к файлу subjects
    :param workbook: объект workbook
    """
    # обнуляем на всякий случай проверку
    ws = workbook.active
    ws.data_validations.dataValidation = []

    # создаем новые значения для прверке данных
    """имена столбцов с индексами"""
    row1 = ws.iter_cols(max_row=1)
    row2 = ws.iter_cols(max_row=2, min_row=2)
    name_row1 = [(str(x[0].value).replace('\n', '_').replace("print_position", "image_position"), x[0].coordinate[0])
                 for x in row1]
    name_row2 = [(str(x[0].value), x[0].coordinate[0]) for x in row2]
    columns = [x if y[0] == 'None' else (f"product_{y[0]}", y[1]) for x, y in zip(name_row1, name_row2)]
    """словарь с именами и столбцами"""
    coordinates = dict(columns)
    for name, column in coordinates.items():
        append_data_validation_subj(name, column, ws)
    workbook.save(file)


def update_templates_create(template, file):
    """
    Функция добавляет название templates, если его уже там нет
    :param template: имя шаблона template
    :param file: путь к Templates2create
    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    if template not in [cell.value for row in ws.iter_rows(min_row=2, min_col=1) for cell in row]:
        for row in ws.iter_rows(min_row=2, min_col=1):
            if row[0].value is None:
                row[0].value = template
                break
        else:
            shape = ws.max_row
            ws[f"A{shape + 1}"] = template
    wb.save(file)
    wb.close()


def get_columns_subj(worksheet):
    """
    Функция парсит файл Subjects и возвращает список имен столбцов в нижним регистре
    каждый продукт будет иметь слово предикат 'product_'
    :param worksheet: - объект worksheet
    """
    first_row = [str(x[0].value).replace('\n', '_') for x in worksheet.iter_cols(max_row=1)]
    second_row = [str(x[0].value) for x in worksheet.iter_cols(max_row=2, min_row=2)]
    col_names = [x.lower() if y == 'None' else f"product_{y.lower()}" for x, y in zip(first_row, second_row)]
    return col_names


def check_subject(data, user_path, user_name):
    """
    Функция проверяет наличие subject в БД И если его не обнаруживает то он добавляется в БД
    :param data: данные openxlsx
    :param user_name: имя юзера
    :return:
    """
    sql = f"""select name from subject where name = '{data['name'].value}'"""
    result = execute_read_query(get_conn(), sql)
    if not result:
        name = data['name'].value
        template = data['template'].value
        author = user_name
        count = len(os.listdir(user_path + f"/{name}/arts"))
        sql = f"""INSERT INTO subject VALUES ('{name}', {count}, '{template}', '{author}', '/subjects/{name}');"""
        execute_query(get_conn(), sql)


def insert_product_group(product, description):
    """
        Функция парсит объкт excel и вставляет данные в бд
        :param description: строка excel (через атрибут value получаем значения)
        :param product: название продукта
        """
    subj = description['name'].value
    common_photo = description['common_photo'].value
    mockup = description['mockup'].value
    infographic = description['infographic'].value
    infographic_color = description['infographic_color'].value
    scale_print = description['scale_print, %'].value
    brand = description['brand'].value
    image_position = description['print_position'].value
    color_mockup = "#" + description['background_mockup'].fill.start_color.index[2:]
    color_print = "#" + description['background_print'].fill.start_color.index[2:]
    backgr_color_mockup = '#' + str(ImageColor.getcolor(color_mockup, "RGB")).replace(",", "")
    backgr_color_print = '#' + str(ImageColor.getcolor(color_print, "RGB")).replace(",", "")

    sql = f"""select name, path_type from product_type where class = '{product}'"""
    varieties = [(item[0], item[1]) for item in execute_read_query(get_conn(), sql)]
    sql = f"""INSERT INTO product_group values"""
    query = [
        f"""('{subj}_{name}', '{subj}', '{name}', '{common_photo}', '{mockup}', '{infographic}', '{infographic_color}', {scale_print}, '{brand}', '{image_position}', '{backgr_color_mockup}', '{backgr_color_print}', '/{subj}{path}')"""
        for name, path in varieties]
    result = sql + ", \n".join(query) + ";"
    try:
        flag = execute_query(get_conn(), result)
    except Exception as ex:
        logger.exception('Error')
    finally:
        return flag


def insert_product(subj, user, product):
    """
    Функция вставляет данные по товару product в таблицу  product
    :param subj: subject
    :param user: путь к рабочей директории
    :param product: конкретный продукт
    """
    subj = subj.value
    # подсчет кол-ва файлов в папке mockups
    count = len(os.listdir(user + f"/{subj}/arts"))
    sql = f"""select name, path_type from product_type where class = '{product}'"""
    varieties = [(item[0], item[1]) for item in execute_read_query(get_conn(), sql)]

    sql = f"""INSERT INTO product values"""
    query = [
        f"""('{subj}_{name}-{num}', '{subj}_{name}', {num}, 0, '/subjects/{subj}{path}/mockups/item_{num}', '/subjects/{subj}{path}/prints/{num}.png')"""
        for name, path in varieties for num in range(count)]
    result = sql + ", \n".join(query) + ";"
    try:
        flag = execute_query(get_conn(), result)
    except Exception as ex:
        logger.exception('Error')
    finally:
        return flag


def transfer_data(user, name, product, storage_path):
    """
    Функция переносит файлы из user/Subjects/name в storage/subjects а также
    проверяет на наличие продуктов и переносит arts если продуктов больше нет
    """
    import shutil
    if os.path.isdir(os.path.join(storage_path, name)):
        if os.path.isdir(os.path.join(storage_path, name, product)):
            shutil.rmtree(os.path.join(storage_path, name, product))
        os.replace(os.path.join(user, name, product), os.path.join(storage_path, name, product))
    else:
        os.makedirs(os.path.join(storage_path, name))
        os.replace(os.path.join(user, name, product), os.path.join(storage_path, name, product))
        shutil.copytree(os.path.join(user, name, 'arts'), os.path.join(storage_path, name, 'arts'))

    if ["arts"] == os.listdir(os.path.join(user, name)):
        if os.path.isdir(os.path.join(storage_path, name, 'arts')):
            shutil.rmtree(os.path.join(user, name))
        logger.info(f"Данные {name} полностью перенесены в storage")


def update_templates_created(path):
    try:
        sql = "select name from template"
        result = (value[0] for value in execute_read_query(get_conn(), sql))
        wb = openpyxl.load_workbook(filename=path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            row[0].value = ''
        for row, value in zip(ws.iter_rows(min_row=2), result):
            row[0].value = value
        wb.save(filename=path)
        wb.close()
    except Exception as ex:
        logger.error(ex)


def main():
    pass


if __name__ == '__main__':
    main()
