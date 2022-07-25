import os
import pandas as pd
import openpyxl
from connect import to_json, get_group, update_template, get_users_templates, path_templatecreate, path_templatecreated, \
    path_storage_templates, path_storage_subjects, execute_query, get_conn, execute_read_query, get_users_subjects, \
    is_correct, template_exist, get_columns_subj, update_templates_create, transfer_data, insert_product, \
    insert_product_group, data_validation_subj, data_validation_temp, path_storage_template_group, check_path, \
    check_subject, update_templates_created
from openpyxl.styles import PatternFill
import logging


def init_logger(name):
    logger = logging.getLogger(name)
    FORMAT = u'%(asctime)s - %(name)s:%(lineno)s - %(levelname)s - %(message)s'
    logger.setLevel(logging.DEBUG)
    sh = logging.StreamHandler()
    sh.setFormatter(logging.Formatter(FORMAT))
    sh.setLevel(logging.DEBUG)
    fh = logging.FileHandler(filename='logs/info.log')
    fh.setFormatter(logging.Formatter(FORMAT))
    fh.setLevel(logging.DEBUG)
    logger.addHandler(sh)
    logger.addHandler(fh)
    logger.debug('logger was initialized')


init_logger('log')
logger = logging.getLogger("log.main")

'-------------------------------------------------------------------------------------------------------------------'


# часть первая работа с template
def check_templates(user_dir, file):
    """
    Проверка сооствествия содержимого файла Templates и директории workdir/templates
    Если в директории отсутствует како-либо шаблон или группа из списка Templates.xlsx,
    то он будет помечен красным цветом
    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    # set_xlsx_file - список файлов в папке Templates
    set_xlsx_file = set(
        [file.rsplit(".", 1)[0] for file in os.listdir(user_dir) if file.endswith('xlsx') and not file.startswith("~")])
    set_png_file = set(
        [file.rsplit(".", 1)[0] for file in os.listdir(user_dir) if file.endswith('png') and not file.startswith("~")])
    try:
        # список templates в файле Templates.xlsx
        set_template = set(pd.read_excel(file)["templates"].values)

        # список group в файле Templates.xlsx
        set_group = set(pd.read_excel(file)["group"].values)

        # symm_diff_templates - это template которые отсутствуют в одном из источников
        symm_diff_templates = set_xlsx_file ^ set_template
        for template in ws.iter_rows(min_row=2, max_col=1):
            if template[0].value in symm_diff_templates:
                template[0].fill = PatternFill('solid', fgColor="FF0000")

        # symm_diff_group - это group которые отсутствуют в одном из источников
        symm_diff_group = set_png_file ^ set_group
        for group in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            sql = f"select name from template_group where name = '{group[0].value}'"
            name_group = execute_read_query(get_conn(), sql)
            if not name_group:
                if group[0].value in symm_diff_group:
                    group[0].fill = PatternFill('solid', fgColor="FF0000")
    except KeyError:
        logger.error(f"Файл {file} должен содержать столбцы 'templates'и 'group'")
    finally:
        wb.save(file)
        wb.close()


def preprocessing_templates(user_path, file_path, storage_templates, storage_template_group, template_created,
                            template_create):
    """
    Сбор всех excel и png файлов их папки Templates. Excel преобразуется в json и перемещается в хранилище
    проверка на напиличе группы template в БД и добавление в БД в случае отсутствия.
    Добавления данных с template и обновление Templates2Created и Templates2Create
    """
    files_excel = [file for file in os.listdir(user_path) if file.endswith("xlsx") and not file.startswith("~")]
    if not files_excel:
        logger.info(f"в папке {user_path} отсутствуют xlsx шаблоны")
    images = [file for file in os.listdir(user_path) if file.endswith("png") and not file.startswith("~")]
    if not images:
        logger.info(f"в папке {user_path} отсутствуют png шаблоны")
    for file in files_excel:  # Преобразуем их в json
        name = file.rsplit('.', 1)[0]
        group = get_group(file_path, name)
        if not group:  # Если не найдено в таблице то мы пропускаем данный template
            continue

        """Проверка наличие group в БД"""
        sql = f"select name from template_group where name = '{group}'"
        name_group = execute_read_query(get_conn(), sql)

        """Если template_group (png) нет в папке и нет в БД то ничего не добавляем т.е пропускаем"""
        if group not in [x.rsplit('.', 1)[0] for x in images] and not name_group:
            logger.info(f"Группа {group} отсутствует в рабочей директории {user_path} и в template_group БД")
            logger.info(f"Template {name} не будет добавлен в storage и БД")
            continue

        if not name_group:  # Проверка группы в template_group
            sql = f"""insert into template_group (name, path_image) values ('{group}', '/template_group/{group}.png')"""  # вставить значения в template_group
            execute_query(get_conn(), sql)

        author = user_path.split('/')[1]
        path_to_excel = f"/templates/{name}.xlsx"
        path_to_json = f"/templates/{name}.json"

        sql = f"INSERT INTO template VALUES ('{name}', '{group}', '{author}', '{path_to_excel}', '{path_to_json}');"
        if execute_query(get_conn(), sql):
            update_template(template=name, templates_xlsx=file_path, template_create=template_create,
                            template_created=template_created)
            logger.info(f"Template {name} группы {group} добавлен в БД")

            """перенос с заменой в случае нахождение таких же файлов"""
            json_name = to_json(user_path + '/' + file)  # Преобразование документ и возврат имени
            os.replace(user_path + '/' + file, os.path.join(storage_templates, file))
            os.replace(user_path + '/' + json_name, os.path.join(storage_templates, json_name))
            logger.info(f"Template {name} перенесен в storage")
            """перенос png если он есть в рабочей директории"""
            if group in [x.rsplit('.', 1)[0] for x in images]:
                os.replace(user_path + '/' + group + '.png', os.path.join(storage_template_group, group + '.png'))
                logger.info(f"Template_group {group} перенесен в storage")
        else:
            logger.info(f"Template {name} группы {group} не добавлен в БД. Возникли проблемы при добавлении данных в "
                        f"таблицу template")
    wb = openpyxl.load_workbook(file_path)
    data_validation_temp(wb, file_path)
    wb.close()


'-------------------------------------------------------------------------------------------------------------------'
# часть вторая работа с subjects


def preprocessing_subj(user, file, user_name, template_create_path, path_storage_subjects):
    """
       Функция обрабатывает subjects
       цикл по каждой строке файла Subjects, проверка директорий (is_correct), наличие template (is_template_exist),
       выбор продукта (check_options); редактирования файла Subjects в случае ошибок, дополнение templatescreate,
       Если ошибок не найдено, функция добавляет данные в product_group и product
       :param user:
       :param file:
       :param template_create_path:
       """
    wb = openpyxl.load_workbook(file)
    data_validation_subj(wb, file)
    ws = wb.active
    columns = get_columns_subj(ws)
    delete_list = []
    for row_number, (col, row) in enumerate(zip(columns, ws.iter_rows(min_row=3))):
        named_data = dict(zip(columns, (x for x in row)))
        # check_dir = is_correct(named_data['name'], user)
        if named_data['name'].value is None:
            continue
        if named_data['name'].value == 'default':
            continue
        check_template = template_exist(named_data['template'])
        products = [k for k in named_data if 'product' in k]
        yes_products = [k.split("_", 1)[1] for k, v in named_data.items() if 'product' in k and v.value == 'yes']
        flags = [v.value for k, v in named_data.items() if 'product' in k]
        check_options = any(True if x == "yes" else False for x in flags)
        check_dir = check_path(named_data['name'], user, yes_products)
        if not check_dir:
            named_data['errors'].value = "Проверьте правильность директорий"
            named_data['name'].fill = PatternFill('solid', fgColor="FF0000")
        if not check_template:
            named_data['template'].fill = PatternFill('solid', fgColor="FFFF00")
            update_templates_create(named_data['template'].value, template_create_path)
        if not check_options:
            for product in products:
                named_data[product].fill = PatternFill('solid', fgColor="FFFF00")
        if all((check_dir, check_template, check_options)):
            """Проверка subject"""
            check_subject(named_data, user_path=user, user_name=user_name)

            for product in products:
                if named_data[product].value == "yes":
                    product_name = product.split("_", 1)[1]
                    subj = named_data['name']
                    check_product_group = insert_product_group(product_name, named_data)
                    check_product = insert_product(subj, user, product_name)
                    if all((check_product, check_product_group)):
                        transfer_data(user, subj.value, product_name, path_storage_subjects)
                        index = row_number + 3
                        if index not in delete_list:
                            delete_list.append(index)
                        logger.info(f"продукт {product_name} {subj.value} добавлен в БД")
                    else:
                        named_data['errors'].value = "Возникла ошибка при добавлении данных в таблицу product_group," \
                                                     "product БД"
                        logger.info(f"Ошибка при добавлении продукта {product_name} {subj.value} в БД")
    for ind, row in enumerate(delete_list):
        ws.delete_rows(row - ind)
    wb.save(file)
    wb.close()


'-------------------------------------------------------------------------------------------------------------------'


def main():
    for user, path in get_users_templates():
        check_templates(user, path)
    for user, path in get_users_templates():
        preprocessing_templates(user, path, path_storage_templates,
                                storage_template_group=path_storage_template_group,
                                template_create=path_templatecreate,
                                template_created=path_templatecreated)
    for user, path, user_name in get_users_subjects():
        preprocessing_subj(user, path, user_name,
                           template_create_path=path_templatecreate,
                           path_storage_subjects=path_storage_subjects)
    update_templates_created(path_templatecreated)


if __name__ == '__main__':
    logger.info("Starting")
    main()
    logger.info("Finishing")
